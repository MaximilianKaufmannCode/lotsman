# SPDX-License-Identifier: BUSL-1.1
# Copyright (c) 2026 Maximilian Kaufmann. See LICENSE (Business Source License 1.1).

"""Two-step xlsx import — Step 1: Preview.

Parses the uploaded xlsx, classifies headers as standard / known-custom / unknown,
stores parsed rows in Redis (gzipped), and returns the preview with session ID.
"""

from __future__ import annotations

import io
import unicodedata
import uuid
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any

from registry_service.application.dto import ColumnInfo, ImportPreviewDTO
from registry_service.application.import_session_codec import dumps_session
from registry_service.application.ports import Clock, DocumentTypeRepository, EventOutbox
from registry_service.domain.errors import RequiredFieldMissingError
from registry_service.domain.events import ImportPreviewStarted

# ---------------------------------------------------------------------------
# Standard header → field mapping (shared with import_xlsx.py)
# ---------------------------------------------------------------------------

# Maps normalised Russian header text → standard field key
_STANDARD_HEADERS: dict[str, str] = {
    "контрагент": "asset",
    "компания": "asset",
    "тип": "type_code",
    "тип документа": "type_code",
    "документ": "type_code",
    "название документа": "type_code",
    "no документа": "doc_number",  # NFKC normalises № → no
    "№ документа": "doc_number",
    "номер документа": "doc_number",
    "действ. до": "expires_at",
    "действует до": "expires_at",
    "дата истечения": "expires_at",
    "срок": "expires_at",
    "ответственный": "responsible_user",
    "ответственный сотрудник": "responsible_user",
    "заметки": "notes",
    "примечание": "notes",
    "комментарий": "notes",
}

_SESSION_TTL = 1800  # 30 minutes


def _normalise(s: str) -> str:
    return unicodedata.normalize("NFKC", s).strip().lower()


def _is_blank(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def _coerce_date(v: Any) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _detect_type(samples: list[Any]) -> str:
    """Auto-detect the best field type from a list of sample values."""
    non_empty = [s for s in samples if not _is_blank(s)]
    if not non_empty:
        return "text"

    # Try date
    if all(_coerce_date(s) is not None for s in non_empty):
        return "date"

    # Try number
    def _is_number(v: Any) -> bool:
        try:
            float(str(v).strip().replace(",", "."))
            return True
        except (ValueError, TypeError):
            return False

    if all(_is_number(s) for s in non_empty):
        return "number"

    return "text"


@dataclass(slots=True)
class ImportXlsxPreview:
    """Use case: parse xlsx, classify headers, store session in Redis."""

    type_repo: DocumentTypeRepository
    outbox: EventOutbox
    clock: Clock
    redis_url: str  # injected from settings

    async def execute(
        self,
        *,
        actor_id: uuid.UUID,
        file_bytes: bytes,
        request_id: str | None = None,
    ) -> ImportPreviewDTO:
        try:
            from openpyxl import load_workbook  # type: ignore[import-untyped]
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for xlsx preview") from exc

        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        ws = wb.active
        if ws is None:
            raise RequiredFieldMissingError("Empty workbook — no active sheet found")

        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row or all(_is_blank(c) for c in header_row):
            raise RequiredFieldMissingError("First row is empty — no headers found")

        # Collect non-blank headers and their column indices
        headers: list[tuple[int, str]] = []  # (col_idx, raw_header)
        for idx, cell in enumerate(header_row):
            if not _is_blank(cell):
                headers.append((idx, str(cell).strip()))

        if not headers:
            raise RequiredFieldMissingError("No headers found in the xlsx file")

        # Read all data rows (up to 10,000 rows for memory safety)
        all_rows: list[tuple[Any, ...]] = []
        for row in rows_iter:
            if row is None or all(_is_blank(c) for c in row):
                continue
            all_rows.append(row)
            if len(all_rows) >= 10_000:
                break

        rows_total = len(all_rows)

        # Collect sample values per column (up to 3 non-blank)
        samples_by_col: dict[int, list[Any]] = {idx: [] for idx, _ in headers}
        for row in all_rows:
            for idx, _ in headers:
                if idx < len(row) and not _is_blank(row[idx]) and len(samples_by_col[idx]) < 3:
                    samples_by_col[idx].append(row[idx])

        # Load all document types for custom schema matching
        all_types = await self.type_repo.list_all()

        # Build a lookup: normalised display_name → (type_code, field_key)
        custom_by_display: dict[str, tuple[str, str]] = {}
        for dt in all_types:
            for cf in dt.custom_field_schema:
                key = _normalise(cf.display_name)
                if key not in custom_by_display:
                    custom_by_display[key] = (dt.code, cf.key)

        # Classify headers
        known_columns: list[ColumnInfo] = []
        unknown_columns: list[ColumnInfo] = []

        for col_idx, raw_header in headers:
            normed = _normalise(raw_header)
            samples = samples_by_col.get(col_idx, [])

            # 1. Standard header?
            std_field = _STANDARD_HEADERS.get(normed)
            if std_field is not None:
                known_columns.append(
                    ColumnInfo(header=raw_header, matched_to=std_field, samples=samples)
                )
                continue

            # 2. Known custom field?
            custom_match = custom_by_display.get(normed)
            if custom_match is not None:
                type_code, field_key = custom_match
                known_columns.append(
                    ColumnInfo(
                        header=raw_header,
                        matched_to=f"custom:{type_code}:{field_key}",
                        samples=samples,
                    )
                )
                continue

            # 3. Unknown — auto-detect suggested type
            suggested = _detect_type(samples)
            unknown_columns.append(
                ColumnInfo(
                    header=raw_header,
                    matched_to="unknown",
                    suggested_type=suggested,
                    samples=samples,
                )
            )

        # Store parsed session in Redis
        session_id = str(uuid.uuid4())
        session_data = {
            "headers": headers,
            "rows": all_rows,
            "known_columns": [(c.header, c.matched_to) for c in known_columns],
            "unknown_headers": [c.header for c in unknown_columns],
        }
        await self._store_session(session_id, session_data)

        # Emit audit event (no values in payload)
        now = self.clock.now()
        event = ImportPreviewStarted(
            rows_total=rows_total,
            unknown_count=len(unknown_columns),
            actor_id=actor_id,
            request_id=request_id,
            occurred_at=now,
        )
        await self.outbox.publish(event.as_envelope(), topic=event.topic)

        return ImportPreviewDTO(
            import_session_id=session_id,
            rows_total=rows_total,
            known_columns=known_columns,
            unknown_columns=unknown_columns,
        )

    async def _store_session(self, session_id: str, data: dict[str, Any]) -> None:
        """Persist session data to Redis as gzipped msgpack (no pickle — CWE-502)."""
        import redis.asyncio as aioredis

        compressed = dumps_session(data)
        async with aioredis.from_url(self.redis_url) as r:  # type: ignore[no-untyped-call]
            await r.set(f"import:session:{session_id}", compressed, ex=_SESSION_TTL)
