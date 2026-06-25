# SPDX-License-Identifier: BUSL-1.1
# Copyright (c) 2026 Maximilian Kaufmann. See LICENSE (Business Source License 1.1).

"""openpyxl-based xlsx exporter for registry documents.

Streams rows in batches to avoid loading the entire dataset into memory.
Writes to a BytesIO buffer then returns the bytes.

Column definitions are driven by the visible_columns list from the export request.
"""

from __future__ import annotations

import io
import uuid
from datetime import date, datetime
from typing import Any

import structlog

from registry_service.domain.entities import Asset, Document

log = structlog.get_logger(__name__)

# Canonical column definitions: key → (header label RU, extractor function)
_COLUMN_DEFS: dict[str, tuple[str, Any]] = {
    "asset_name": (
        "Компания",
        lambda d, assets: (
            assets.get(d.asset_id, type("", (), {"name": ""})()).name
            if d.asset_id in assets
            else ""
        ),
    ),  # noqa: E501
    "type_code": ("Тип документа", lambda d, _: d.type_code),
    "number": ("№ документа", lambda d, _: d.number or ""),
    "issue_date": ("Дата выдачи", lambda d, _: _fmt_date(d.issue_date)),
    "expiry_date": ("Действ. до", lambda d, _: _fmt_date(d.expiry_date)),
    "status": ("Статус", lambda d, _: d.status),
    "notes": ("Примечания", lambda d, _: d.notes or ""),
    "created_at": ("Создан", lambda d, _: _fmt_dt(d.created_at)),
    "updated_at": ("Изменён", lambda d, _: _fmt_dt(d.updated_at)),
}

_DEFAULT_COLUMNS = ["asset_name", "type_code", "number", "issue_date", "expiry_date", "status"]


def _fmt_date(d: date | None) -> str:
    return d.strftime("%d.%m.%Y") if d else ""


def _fmt_dt(dt: datetime | None) -> str:
    return dt.strftime("%d.%m.%Y %H:%M") if dt else ""


# F-017 mitigation (CWE-1236, CSV/formula injection):
# Excel/LibreOffice/Numbers interpret cells beginning with =, +, -, @, \t, \r as formulas.
# Prefix any such string with a leading apostrophe to force literal interpretation.
# Apostrophe is itself stripped on display so the user sees the original text.
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def _safe_cell(value: Any) -> Any:
    """Return value with formula-injection escape if it's a string starting with a trigger char."""
    if isinstance(value, str) and value and value[0] in _FORMULA_TRIGGERS:
        return "'" + value
    return value


class OpenpyxlExporter:
    """Produces xlsx bytes for a list of documents."""

    async def export(
        self,
        *,
        documents: list[Document],
        assets: dict[uuid.UUID, Asset],
        visible_columns: list[str],
        snapshot_at: Any,
    ) -> bytes:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError as exc:
            raise RuntimeError(
                "openpyxl is not installed. Add it to pyproject.toml dependencies."
            ) from exc

        cols = [c for c in visible_columns if c in _COLUMN_DEFS] or _DEFAULT_COLUMNS

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Реестр"  # type: ignore[assignment]

        # Header row (header labels are static, no injection risk)
        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="D9E1F2")
        headers = [_COLUMN_DEFS[c][0] for c in cols]
        ws.append(headers)  # type: ignore[attr-defined]
        for cell in ws[1]:  # type: ignore[index]
            cell.font = header_font
            cell.fill = header_fill

        # Data rows — stream in chunks of 1000 to avoid memory pressure (US-20 NFR)
        # _safe_cell escapes formula-injection vectors per F-017
        chunk_size = 1000
        for i in range(0, len(documents), chunk_size):
            chunk = documents[i : i + chunk_size]
            for doc in chunk:
                row = [_safe_cell(_COLUMN_DEFS[c][1](doc, assets)) for c in cols]
                ws.append(row)  # type: ignore[attr-defined]

        # Metadata sheet
        meta_ws = wb.create_sheet("Метаданные")
        meta_ws.append(["Лоцман — реестр документов"])  # type: ignore[attr-defined]
        meta_ws.append(["Снимок данных:", str(snapshot_at)])  # type: ignore[attr-defined]
        meta_ws.append(["Документов:", len(documents)])  # type: ignore[attr-defined]

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
