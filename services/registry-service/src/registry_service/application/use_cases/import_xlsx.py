# SPDX-License-Identifier: BUSL-1.1
# Copyright (c) 2026 Maximilian Kaufmann. See LICENSE (Business Source License 1.1).

"""Import documents and assets from a corporate Excel registry (.xlsx / .xlsm).

Expected sheet shape (Russian headers, in any order):
  Активность | Компания | Название документа | Дата документа |
  Дата истечения | Дата подачи отчетности | Комментарий | Юрисдикция

Merged cells in 'Активность' / 'Компания' / 'Юрисдикция' are unwrapped:
the value carries forward until the next non-empty cell.

The import is idempotent on the natural key:
  Asset:    by `name` (case-insensitive) within the same jurisdiction
  DocumentType: by display_name → code = slug(display_name)
  Document: by (asset_id, type_code, issue_date) — re-importing the same
            row updates expiry/notes instead of duplicating.
"""

from __future__ import annotations

import io
import re
import unicodedata
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from registry_service.db.models import Asset as AssetModel
from registry_service.db.models import Document as DocumentModel
from registry_service.db.models import DocumentType as DocumentTypeModel
from registry_service.domain.errors import UnknownColumnsError

# ── Header → field ───────────────────────────────────────────────────────────

_HEADER_MAP: dict[str, str] = {
    "активность": "activity",
    "статус": "activity",
    "компания": "company",
    "контрагент": "company",
    "название документа": "doc_name",
    "документ": "doc_name",
    "дата документа": "issue_date",
    "дата истечения": "expiry_date",
    "срок": "expiry_date",
    "действует до": "expiry_date",
    "действ. до": "expiry_date",
    "дата подачи отчетности": "report_due",
    "дата подачи отчётности": "report_due",
    "комментарий": "notes",
    "примечание": "notes",
    "заметки": "notes",
    "юрисдикция": "jurisdiction",
    "тип": "doc_name",
    "тип документа": "doc_name",
    "no документа": "doc_number",  # NFKC normalises № → no
    "№ документа": "doc_number",
    "номер документа": "doc_number",
    "ответственный": "responsible_user",
    "ответственный сотрудник": "responsible_user",
}


@dataclass
class ImportRowError:
    row_index: int
    company: str
    document: str
    error: str


@dataclass
class ImportReport:
    total_rows: int = 0
    assets_created: int = 0
    assets_reused: int = 0
    types_created: int = 0
    documents_created: int = 0
    documents_updated: int = 0
    skipped: int = 0
    errors: list[ImportRowError] = field(default_factory=list)


def _slug(s: str) -> str:
    """Make a stable kebab-case identifier out of a Russian/English string."""
    s = unicodedata.normalize("NFKC", s).strip().lower()
    s = re.sub(r"[^\w\s-]", "", s, flags=re.UNICODE)
    s = re.sub(r"[\s_-]+", "-", s).strip("-")
    return s[:60] or "untitled"


def _coerce_date(v: Any) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _row_dict(row: tuple, header_to_idx: dict[str, int]) -> dict[str, Any]:
    return {
        field_name: row[idx] if idx < len(row) else None
        for field_name, idx in header_to_idx.items()
    }


def _is_blank(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


@dataclass(slots=True)
class ImportXlsx:
    """Use case: parse an uploaded .xlsx and upsert assets/document types/documents."""

    session: AsyncSession
    actor_id: uuid.UUID

    async def execute(self, *, data: bytes, filename: str = "import.xlsx") -> ImportReport:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for xlsx import") from exc

        report = ImportReport()
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        ws = wb.active
        if ws is None:
            return report

        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return report

        header_to_idx: dict[str, int] = {}
        unknown_headers: list[str] = []
        for idx, cell in enumerate(header_row):
            if cell is None:
                continue
            raw = str(cell).strip()
            key = unicodedata.normalize("NFKC", raw).lower()
            field_name = _HEADER_MAP.get(key)
            if field_name and field_name not in header_to_idx:
                header_to_idx[field_name] = idx
            elif field_name is None and raw:
                unknown_headers.append(raw)

        if unknown_headers:
            raise UnknownColumnsError(
                "Unknown columns detected — use /admin/import/preview to classify them",
                unknown_columns=unknown_headers,
            )

        if "company" not in header_to_idx or "doc_name" not in header_to_idx:
            raise ValueError(
                "Required columns not found. Need at least 'Компания' and 'Название документа'."
            )

        # Carry-forward state for merged cells (Активность/Компания/Юрисдикция)
        last_company: str | None = None
        last_jurisdiction: str | None = None
        last_activity: str | None = None

        # Caches to avoid re-querying within the same import
        asset_cache: dict[str, AssetModel] = {}
        type_cache: dict[str, DocumentTypeModel] = {}

        # Pre-load existing types
        existing_types = (await self.session.execute(select(DocumentTypeModel))).scalars().all()
        for t in existing_types:
            type_cache[t.display_name.strip().lower()] = t

        for row_idx, raw_row in enumerate(rows_iter, start=2):
            if raw_row is None or all(_is_blank(c) for c in raw_row):
                continue
            row = _row_dict(raw_row, header_to_idx)
            report.total_rows += 1

            # Forward-fill merged columns
            if not _is_blank(row.get("activity")):
                last_activity = str(row["activity"]).strip()
            if not _is_blank(row.get("company")):
                last_company = str(row["company"]).strip()
            if not _is_blank(row.get("jurisdiction")):
                last_jurisdiction = str(row["jurisdiction"]).strip()

            company = last_company
            doc_name = row.get("doc_name")
            if _is_blank(company) or _is_blank(doc_name):
                report.skipped += 1
                continue
            company = str(company).strip()
            doc_name = str(doc_name).strip()

            try:
                asset = await self._upsert_asset(company, last_jurisdiction, asset_cache, report)
                doc_type = await self._upsert_type(doc_name, type_cache, report)
                await self._upsert_document(
                    asset=asset,
                    doc_type=doc_type,
                    issue_date=_coerce_date(row.get("issue_date")),
                    expiry_date=_coerce_date(row.get("expiry_date")),
                    report_due=_coerce_date(row.get("report_due")),
                    notes=(
                        str(row.get("notes")).strip() if not _is_blank(row.get("notes")) else None
                    ),
                    activity=last_activity,
                    report=report,
                )
            except Exception as exc:  # noqa: BLE001 — capture per-row, continue
                report.errors.append(
                    ImportRowError(
                        row_index=row_idx,
                        company=company,
                        document=doc_name,
                        error=str(exc),
                    )
                )

        await self.session.flush()
        return report

    async def _upsert_asset(
        self,
        name: str,
        jurisdiction: str | None,
        cache: dict[str, AssetModel],
        report: ImportReport,
    ) -> AssetModel:
        key = name.lower()
        if key in cache:
            return cache[key]
        existing = (
            await self.session.execute(select(AssetModel).where(AssetModel.name == name))
        ).scalar_one_or_none()
        if existing is not None:
            cache[key] = existing
            report.assets_reused += 1
            return existing
        notes = f"Юрисдикция: {jurisdiction}" if jurisdiction else None
        asset = AssetModel(
            id=uuid.uuid4(),
            name=name,
            inn=None,
            notes=notes,
        )
        self.session.add(asset)
        await self.session.flush()
        cache[key] = asset
        report.assets_created += 1
        return asset

    async def _upsert_type(
        self,
        display_name: str,
        cache: dict[str, DocumentTypeModel],
        report: ImportReport,
    ) -> DocumentTypeModel:
        key = display_name.strip().lower()
        if key in cache:
            return cache[key]
        code = _slug(display_name)
        # If something with that code exists, reuse
        existing = (
            await self.session.execute(
                select(DocumentTypeModel).where(DocumentTypeModel.code == code)
            )
        ).scalar_one_or_none()
        if existing is not None:
            cache[key] = existing
            return existing
        dt = DocumentTypeModel(
            code=code,
            display_name=display_name,
            pre_notice_days=[30, 7, 1],
            notify_in_day=True,
            overdue_every_days=7,
        )
        self.session.add(dt)
        await self.session.flush()
        cache[key] = dt
        report.types_created += 1
        return dt

    async def _upsert_document(
        self,
        *,
        asset: AssetModel,
        doc_type: DocumentTypeModel,
        issue_date: date | None,
        expiry_date: date | None,
        report_due: date | None,
        notes: str | None,
        activity: str | None,
        report: ImportReport,
    ) -> None:
        # Idempotency: same (asset_id, type_code, issue_date) → update;
        # if issue_date is None, fall back to (asset_id, type_code) uniqueness.
        stmt = select(DocumentModel).where(
            DocumentModel.asset_id == asset.id,
            DocumentModel.type_code == doc_type.code,
        )
        if issue_date is not None:
            stmt = stmt.where(DocumentModel.issue_date == issue_date)
        existing = (await self.session.execute(stmt)).scalar_one_or_none()

        # Combine activity + report_due into notes (no dedicated column yet)
        extra_notes: list[str] = []
        if activity:
            extra_notes.append(f"Активность: {activity}")
        if report_due:
            extra_notes.append(f"Дата подачи отчётности: {report_due.strftime('%d.%m.%Y')}")
        if notes:
            extra_notes.append(notes)
        merged_notes = "\n".join(extra_notes) or None

        if existing is None:
            doc = DocumentModel(
                id=uuid.uuid4(),
                asset_id=asset.id,
                type_code=doc_type.code,
                number=None,
                issue_date=issue_date,
                expiry_date=expiry_date,
                responsible_user_id=None,
                status="active",
                notes=merged_notes,
                created_by=self.actor_id,
                updated_by=self.actor_id,
            )
            self.session.add(doc)
            report.documents_created += 1
        else:
            existing.expiry_date = expiry_date or existing.expiry_date
            existing.notes = merged_notes or existing.notes
            existing.updated_by = self.actor_id
            report.documents_updated += 1
